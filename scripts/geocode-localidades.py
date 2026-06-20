#!/usr/bin/env python3
# Geolocaliza la presidencial 2025 (1ª vuelta, acta por mesa del OEP) y emite el
# scatter del mapa en 4 niveles de granularidad + los datos para la BDR.
#
#   pip install openpyxl   # (en un venv)
#   curl -o /tmp/BO.zip https://download.geonames.org/export/dump/BO.zip && unzip /tmp/BO.zip BO.txt -d /tmp
#   python3 scripts/geocode-localidades.py  <ruta_xlsx>  /tmp/BO.txt
#
# Geocode: GeoNames Bolivia (CC-BY) por nombre de localidad, desambiguado por cercanía
# al centroide del departamento; fallback municipio -> centroide de localidades hermanas
# -> provincia/depto + jitter. Coordenadas APROXIMADAS. Ignora Votos0/Voto11/Voto12.
# Salidas: src/data/localidades/{dept,prov,muni,loc}.ts (lazy-load, default export) y
#          db/localidades2025.json (BDR, nivel localidad).

import json, unicodedata, re, hashlib, sys, os
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    '~/Desktop/Elecciones/Elecciones_Generales_BOL1era.xlsx')
GEON = sys.argv[2] if len(sys.argv) > 2 else '/tmp/BO.txt'

PARTIES = ['AP', 'LYP AND', 'APB SUMATE', 'LIBRE', 'FP', 'MAS-IPSP', 'MORENA', 'UNIDAD', 'PDC']
IDX = {'AP': 18, 'LYP AND': 19, 'APB SUMATE': 20, 'LIBRE': 22, 'FP': 23,
       'MAS-IPSP': 24, 'MORENA': 25, 'UNIDAD': 26, 'PDC': 27}  # 21/28/29 = ignorados
CEN = {'La Paz': (-68.0, -15.5), 'Santa Cruz': (-61.8, -17.0), 'Cochabamba': (-65.8, -17.4),
       'Oruro': (-67.6, -18.7), 'Potosí': (-66.5, -20.6), 'Tarija': (-64.1, -21.7),
       'Chuquisaca': (-64.3, -19.9), 'Beni': (-65.0, -14.2), 'Pando': (-67.6, -11.4)}


def norm(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFD', str(s))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9]+', ' ', s.lower())).strip()


def num(x):
    try:
        return int(x)
    except Exception:
        return 0


def jit(cod, amp):
    h = int(hashlib.md5(str(cod).encode()).hexdigest(), 16)
    return ((h % 1000) / 1000 - 0.5) * 2 * amp, (((h // 1000) % 1000) / 1000 - 0.5) * 2 * amp


# 1) agregar el Excel por localidad (con votos por partido)
from openpyxl import load_workbook
ws = load_workbook(XLSX, read_only=True, data_only=True)['Sheet2']
loc = {}
it = ws.iter_rows(values_only=True)
next(it)
for r in it:
    a = loc.setdefault(r[12], {'dep': r[5], 'prov': r[9], 'muni': r[11], 'loc': r[13],
                               'v': {k: 0 for k in PARTIES}})
    for k in PARTIES:
        a['v'][k] += num(r[IDX[k]])
locs = []
for cod, a in loc.items():
    tv = sum(a['v'].values())
    if tv <= 0:
        continue
    a.update(cod=cod, win=max(PARTIES, key=lambda k: a['v'][k]), tot=tv)
    locs.append(a)

# 2) índice GeoNames + geocode (igual que la versión previa)
idx = {}
for line in open(GEON, encoding='utf-8'):
    f = line.rstrip('\n').split('\t')
    if len(f) < 15 or f[6] not in ('P', 'A'):
        continue
    try:
        lat, lon = float(f[4]), float(f[5])
    except Exception:
        continue
    for nm in set([f[1], f[2]] + (f[3].split(',') if f[3] else [])):
        k = norm(nm)
        if k:
            idx.setdefault(k, []).append((lat, lon))


def best(name, dep, maxd):
    c = idx.get(norm(name))
    if not c:
        return None
    cx, cy = CEN[dep]
    bb, bd = None, 1e9
    for lat, lon in c:
        d = ((lon - cx) ** 2 + (lat - cy) ** 2) ** 0.5
        if d < bd:
            bd, bb = d, (lon, lat)
    return bb if bd <= maxd else None


P1 = {a['cod']: (best(a['loc'], a['dep'], 2.0), 'loc') for a in locs}
msum, psum = defaultdict(lambda: [0, 0, 0]), defaultdict(lambda: [0, 0, 0])
for a in locs:
    p, _ = P1[a['cod']]
    if p:
        for agg, key in ((msum, (a['dep'], a['prov'], a['muni'])), (psum, (a['dep'], a['prov']))):
            agg[key][0] += p[0]; agg[key][1] += p[1]; agg[key][2] += 1
mcen = {k: (v[0] / v[2], v[1] / v[2]) for k, v in msum.items()}
pcen = {k: (v[0] / v[2], v[1] / v[2]) for k, v in psum.items()}

out, stat = [], defaultdict(int)
for a in locs:
    p, lvl = P1[a['cod']]
    if not p:
        p = best(a['muni'], a['dep'], 1.5)
        if p:
            lvl = 'muni'; dx, dy = jit(a['cod'], 0.06); p = (p[0] + dx, p[1] + dy)
    if not p:
        mk, pk = (a['dep'], a['prov'], a['muni']), (a['dep'], a['prov'])
        base, lvl, amp = (mcen[mk], 'muni~', 0.06) if mk in mcen else \
            (pcen[pk], 'prov~', 0.12) if pk in pcen else (CEN[a['dep']], 'dept', 0.25)
        dx, dy = jit(a['cod'], amp); p = (base[0] + dx, base[1] + dy)
    stat[lvl] += 1
    out.append({'lon': round(p[0], 3), 'lat': round(p[1], 3), 'p': PARTIES.index(a['win']),
                't': a['tot'], 'n': a['loc'], 'dep': a['dep'], 'prov': a['prov'],
                'muni': a['muni'], 'win': a['win']})

# 3) agregaciones (votos por partido sumados; centroide ponderado por votos)
def aggregate(keyfn, namefn):
    agg = {}
    for i, a in enumerate(locs):
        o = out[i]
        g = agg.setdefault(keyfn(a), {'v': defaultdict(int), 'wx': 0, 'wy': 0, 'wt': 0, 'n': namefn(a)})
        for k in PARTIES:
            g['v'][k] += a['v'][k]
        g['wx'] += o['lon'] * o['t']; g['wy'] += o['lat'] * o['t']; g['wt'] += o['t']
    res = []
    for g in agg.values():
        if g['wt'] <= 0:
            continue
        win = max(PARTIES, key=lambda k: g['v'][k])
        res.append({'lon': round(g['wx'] / g['wt'], 3), 'lat': round(g['wy'] / g['wt'], 3),
                    'p': PARTIES.index(win), 't': sum(g['v'].values()), 'n': g['n']})
    return res

LEVELS = {
    'dept': aggregate(lambda a: a['dep'], lambda a: a['dep']),
    'prov': aggregate(lambda a: (a['dep'], a['prov']), lambda a: a['prov']),
    'muni': aggregate(lambda a: (a['dep'], a['prov'], a['muni']), lambda a: a['muni']),
    'loc': [{'lon': o['lon'], 'lat': o['lat'], 'p': o['p'], 't': o['t'], 'n': o['n']} for o in out],
}

# 4) emitir módulos (lazy-load) + BDR json
os.makedirs(os.path.join(ROOT, 'src/data/localidades'), exist_ok=True)
for name, items in LEVELS.items():
    rows = ','.join('{lon:%s,lat:%s,p:%d,t:%d,n:%s}' % (
        o['lon'], o['lat'], o['p'], o['t'], json.dumps(o['n'], ensure_ascii=False)) for o in items)
    ts = ("// AUTO-GENERADO por scripts/geocode-localidades.py — no editar a mano.\n"
          "// Presidencial 2025 (1ª vuelta, OEP) geolocalizada con GeoNames (CC-BY). Nivel: %s.\n"
          "// p = índice de partido ganador; t = votos válidos; n = nombre. Coords APROXIMADAS.\n"
          "import type { LocPunto } from '../../types';\n"
          "const D: LocPunto[] = [%s];\nexport default D;\n") % (name, rows)
    open(os.path.join(ROOT, 'src/data/localidades', name + '.ts'), 'w').write(ts)
json.dump(out, open(os.path.join(ROOT, 'db/localidades2025.json'), 'w'), ensure_ascii=False)

real = sum(stat[k] for k in ('loc', 'muni', 'muni~'))
print('localidades=%d (loc/municipio %.1f%%, aprox %.1f%%) | dept=%d prov=%d muni=%d loc=%d'
      % (len(out), 100 * real / len(out), 100 * (len(out) - real) / len(out),
         len(LEVELS['dept']), len(LEVELS['prov']), len(LEVELS['muni']), len(LEVELS['loc'])))
print('->  src/data/localidades/{dept,prov,muni,loc}.ts  +  db/localidades2025.json')
